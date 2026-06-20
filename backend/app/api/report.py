from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from datetime import datetime
import os
from typing import List
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import csv

from app.core.database import get_db
from app.api.auth import get_current_user
from app.models.user import User
from app.models.domain import Domain
from app.models.project import Project
from app.models.asset import Asset
from app.models.tara_run import TaraRun
from app.models.tara_step import TaraStep

router = APIRouter(prefix="/reports", tags=["报告导出与脱敏服务"])

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
EXPORTS_DIR = os.path.join(PROJECT_ROOT, "exports_local")

# 建立导出文件夹
if not os.path.exists(EXPORTS_DIR):
    os.makedirs(EXPORTS_DIR)

def get_asset_sn(asset) -> str:
    prefix_map = {
        "hardware": "H",
        "software": "S",
        "data": "D",
        "communication": "C"
    }
    t = str(asset.asset_type).lower().strip()
    prefix = prefix_map.get(t, "A")
    return f"{prefix}-001_{asset.id:03d}"

def build_id_maps(steps: List[TaraStep]):
    goal_set = set()
    claim_set = set()
    control_set = set()
    req_set = set()

    for step in steps:
        out = step.analysis_result.get("final_output", {}) if step.analysis_result else {}
        if step.stage == "stage4":
            decisions = out.get("risk_decisions") or []
            if not decisions and (out.get("cybersecurity_goal") or out.get("cybersecurity_claim")):
                decisions = [{
                    "risk_treatment": out.get("risk_decision", "Retain"),
                    "cybersecurity_goal": out.get("cybersecurity_goal", ""),
                    "cybersecurity_claim": out.get("cybersecurity_claim", "")
                }]
            for rd in decisions:
                rt = str(rd.get("risk_treatment") or rd.get("risk_decision") or "Retain").lower()
                is_exempted = rt in ["accept", "transfer", "share", "retain"]
                
                if is_exempted:
                    claim = str(rd.get("cybersecurity_claim") or "").strip()
                    if claim and claim != "N/A":
                        claim_set.add(claim)
                else:
                    goal = str(rd.get("cybersecurity_goal") or "").strip()
                    if goal and goal != "N/A":
                        goal_set.add(goal)
        elif step.stage == "stage5":
            reqs = out.get("requirements") or []
            if not reqs:
                cso_val = out.get("cso") or "N/A"
                csr_list = out.get("csr") or []
                if cso_val != "N/A" or csr_list:
                    if isinstance(csr_list, list) and csr_list:
                        reqs = [{"cybersecurity_control": cso_val, "cybersecurity_requirement": text} for text in csr_list]
                    else:
                        reqs = [{"cybersecurity_control": cso_val, "cybersecurity_requirement": str(csr_list)}]
            
            for req in reqs:
                ctrl = str(req.get("cybersecurity_control") or "").strip()
                if ctrl and ctrl != "N/A":
                    control_set.add(ctrl)
                rq = str(req.get("cybersecurity_requirement") or "").strip()
                if rq and rq != "N/A":
                    for line in rq.split("\n"):
                        import re
                        cleaned = re.sub(r'^\(\d+\)\s*', '', line).strip()
                        if cleaned and cleaned != "N/A":
                            req_set.add(cleaned)
                            
            summarized_csrs = out.get("summarized_csrs") or []
            for csr in summarized_csrs:
                rq = str(csr.get("cybersecurity_requirement") or "").strip()
                if rq and rq != "N/A":
                    for line in rq.split("\n"):
                        import re
                        cleaned = re.sub(r'^\(\d+\)\s*', '', line).strip()
                        if cleaned and cleaned != "N/A":
                            req_set.add(cleaned)

    sorted_goals = sorted(list(goal_set))
    sorted_claims = sorted(list(claim_set))
    sorted_controls = sorted(list(control_set))
    sorted_reqs = sorted(list(req_set))

    goal_map = {g: f"CSO-{i+1:04d}" for i, g in enumerate(sorted_goals)}
    claim_map = {c: f"CLM-{i+1:04d}" for i, c in enumerate(sorted_claims)}
    control_map = {ct: f"CSC-{i+1:04d}" for i, ct in enumerate(sorted_controls)}
    req_map = {r: f"CSR-{i+1:04d}" for i, r in enumerate(sorted_reqs)}

    return goal_map, claim_map, control_map, req_map

def collect_report_data(steps: List[TaraStep], assets: List[Asset], desensitize: bool) -> List[dict]:
    """
    收集并平铺 TARA 各阶段分析结果数据，返回扁平化字典列表
    """
    steps_by_asset = {}
    for step in steps:
        if step.asset_id not in steps_by_asset:
            steps_by_asset[step.asset_id] = {}
        steps_by_asset[step.asset_id][step.stage] = step
        
    attr_cols = ["Authenticity", "Integrity", "Non-repudiation", "Confidentiality", "Availability", "Authorization", "Privacy"]
    chinese_suffix_map = {
        "Authenticity": "真实性",
        "Integrity": "完整性",
        "Non-repudiation": "抗抵赖性",
        "Confidentiality": "机密性",
        "Availability": "可用性",
        "Authorization": "授权性",
        "Privacy": "隐私性"
    }
    
    row_num = 1
    rows = []
    
    # 对资产按 ID 排序后依次展开子树
    for asset in sorted(assets, key=lambda x: x.id):
        asset_steps = steps_by_asset.get(asset.id, {})
        if not asset_steps:
            continue
            
        stage1_out = asset_steps.get("stage1").analysis_result.get("final_output", {}) if asset_steps.get("stage1") else {}
        stage2_out = asset_steps.get("stage2").analysis_result.get("final_output", {}) if asset_steps.get("stage2") else {}
        stage3_out = asset_steps.get("stage3").analysis_result.get("final_output", {}) if asset_steps.get("stage3") else {}
        stage4_out = asset_steps.get("stage4").analysis_result.get("final_output", {}) if asset_steps.get("stage4") else {}
        stage5_out = asset_steps.get("stage5").analysis_result.get("final_output", {}) if asset_steps.get("stage5") else {}
        
        # 提取高相关属性
        selected_attrs = stage1_out.get("selected_attributes", [])
        if not selected_attrs:
            attrs_dict = stage1_out.get("attributes", {})
            selected_attrs = [k for k, v in attrs_dict.items() if int(v or 0) > 2]
        if not selected_attrs:
            selected_attrs = []
            for attr in attr_cols:
                if stage1_out.get(attr.lower()) not in [None, "None", "Low"]:
                    selected_attrs.append(attr)
        if not selected_attrs:
            selected_attrs = ["Confidentiality", "Integrity", "Availability"]
            
        for attr in selected_attrs:
            # 损害场景
            damage_scenarios = stage2_out.get("damage_scenarios", [])
            if not damage_scenarios:
                if stage2_out.get("damage_scenario"):
                    damage_scenarios = [{
                        "attribute": attr,
                        "damage_scenario_sn": "DS_00001",
                        "damage_scenario": stage2_out.get("damage_scenario"),
                        "impact_rating": stage2_out.get("impact_rating", {}),
                        "overall_impact": stage2_out.get("overall_impact", "Medium")
                    }]
            
            matching_ds = [ds for ds in damage_scenarios if ds.get("attribute") == attr]
            if not matching_ds:
                matching_ds = [{
                    "attribute": attr,
                    "damage_scenario_sn": "DS_N/A",
                    "damage_scenario": "N/A",
                    "impact_rating": {"safety": "Negligible", "financial": "Negligible", "operational": "Negligible", "privacy": "Negligible"},
                    "overall_impact": "Negligible"
                }]
                
            for ds in matching_ds:
                # 威胁场景
                threat_scenarios = stage3_out.get("threat_scenarios", [])
                if not threat_scenarios:
                    if stage3_out.get("threat_scenario"):
                        threat_scenarios = [{
                            "attribute": attr,
                            "damage_scenario_sn": ds.get("damage_scenario_sn"),
                            "threat_id": "TS_00001",
                            "threat_scenario": stage3_out.get("threat_scenario"),
                            "attack_paths": stage3_out.get("attack_paths", []),
                            "final_feasibility": stage3_out.get("final_feasibility", "Medium")
                        }]
                
                matching_ts = [ts for ts in threat_scenarios if ts.get("attribute") == attr and ts.get("damage_scenario_sn") == ds.get("damage_scenario_sn")]
                if not matching_ts:
                    matching_ts = [{
                        "attribute": attr,
                        "damage_scenario_sn": ds.get("damage_scenario_sn"),
                        "threat_id": "TS_N/A",
                        "threat_scenario": "N/A",
                        "attack_paths": [],
                        "final_feasibility": "Medium"
                    }]
                    
                for ts in matching_ts:
                    # 风险处理决策
                    risk_decisions = stage4_out.get("risk_decisions", [])
                    rd = None
                    for r_dec in risk_decisions:
                        if r_dec.get("threat_id") == ts.get("threat_id"):
                            rd = r_dec
                            break
                    if not rd:
                        for r_dec in risk_decisions:
                            if r_dec.get("attribute") == attr:
                                rd = r_dec
                                break
                    if not rd:
                        rd = {
                            "risk_value": stage4_out.get("risk_rating", 1),
                            "risk_treatment": stage4_out.get("risk_decision", "Retain"),
                            "item_change": "",
                            "cybersecurity_goal": "",
                            "cybersecurity_claim": ""
                        }
                        
                    # 网络安全控制与要求
                    requirements = stage5_out.get("requirements", [])
                    matching_reqs = [r for r in requirements if r.get("threat_id") == ts.get("threat_id")]
                    if not matching_reqs:
                        # 兜底：如果为空，但顶级存在 cso 或是 csr，代表这可能是手工录入/人工修改的结构
                        cso_val = stage5_out.get("cso") or "N/A"
                        csr_list = stage5_out.get("csr") or []
                        if cso_val != "N/A" or csr_list:
                            if isinstance(csr_list, list) and csr_list:
                                matching_reqs = []
                                for idx, csr_text in enumerate(csr_list):
                                    matching_reqs.append({
                                        "cybersecurity_control": cso_val if idx == 0 else "N/A",
                                        "allocated_to_device": "yes",
                                        "cybersecurity_requirement": csr_text
                                    })
                            else:
                                csr_val = str(csr_list) if csr_list else "N/A"
                                matching_reqs = [{
                                    "cybersecurity_control": cso_val,
                                    "allocated_to_device": "yes",
                                    "cybersecurity_requirement": csr_val
                                }]
                        else:
                            matching_reqs = [{
                                "cybersecurity_control_id": "N/A",
                                "cybersecurity_control": "N/A",
                                "allocated_to_device": "No",
                                "cybersecurity_requirement_id": "N/A",
                                "cybersecurity_requirement": "N/A"
                            }]
                        
                    # 攻击路径
                    attack_paths = ts.get("attack_paths", [])
                    if not attack_paths:
                        if stage3_out.get("attack_paths"):
                            attack_paths = stage3_out.get("attack_paths")
                        else:
                            attack_paths = [{
                                "attack_path": "N/A",
                                "time_consuming": "no_more_than_1w",
                                "expertise": "proficient",
                                "knowledge_about_toe": "restricted",
                                "window_of_opportunity": "easy",
                                "equipment": "standard",
                                "difficulty": 9,
                                "feasibility": ts.get("final_feasibility", "Medium")
                            }]
                            
                    # Consolidated Requirements (Stage 5) for the columns
                    rt = rd.get("risk_treatment", "Retain")
                    rt_map = {
                        "avoid": "Avoid",
                        "reduce": "Reduce",
                        "mitigate": "Reduce",
                        "share": "Share",
                        "transfer": "Share",
                        "retain": "Retain",
                        "accept": "Retain"
                    }
                    rt_disp = rt_map.get(str(rt).lower(), rt)
                    
                    cybersecurity_control_id_val = "N/A"
                    cybersecurity_control_val = "N/A"
                    allocated_to_device_val = "No"
                    cybersecurity_requirement_id_val = "N/A"
                    cybersecurity_requirement_val = "N/A"
                    
                    if rt_disp == "Reduce":
                        control_ids = []
                        controls = []
                        allocs = []
                        req_ids = []
                        reqs = []
                        for r_idx, req in enumerate(matching_reqs):
                            c_id_val = req.get("cybersecurity_control_id") or "N/A"
                            c_val = req.get("cybersecurity_control") or "N/A"
                            a_val = "Yes" if str(req.get("allocated_to_device", "No")).lower() in ["yes", "true"] else "No"
                            rq_id_val = req.get("cybersecurity_requirement_id") or "N/A"
                            rq_val = req.get("cybersecurity_requirement") or "N/A"
                            if len(matching_reqs) > 1:
                                control_ids.append(f"({r_idx+1}) {c_id_val}")
                                controls.append(f"({r_idx+1}) {c_val}")
                                allocs.append(f"({r_idx+1}) {a_val}")
                                req_ids.append(f"({r_idx+1}) {rq_id_val}")
                                reqs.append(f"({r_idx+1}) {rq_val}")
                            else:
                                control_ids.append(c_id_val)
                                controls.append(c_val)
                                allocs.append(a_val)
                                req_ids.append(rq_id_val)
                                reqs.append(rq_val)
                        cybersecurity_control_id_val = "\n".join(control_ids)
                        cybersecurity_control_val = "\n".join(controls)
                        allocated_to_device_val = "\n".join(allocs)
                        cybersecurity_requirement_id_val = "\n".join(req_ids)
                        cybersecurity_requirement_val = "\n".join(reqs)
                    elif rt_disp == "Avoid":
                        cybersecurity_requirement_val = rd.get("item_change") or "N/A"
                        
                    for ap in attack_paths:
                        row_data = {}
                        row_data["number"] = f"T_{row_num:04d}"
                        row_data["asset_sn"] = get_asset_sn(asset)
                        row_data["asset_name"] = asset.name
                        
                        for c in attr_cols:
                            row_data[c] = "Yes" if c == attr else None
                            
                        row_data["attribute_result"] = f"{chinese_suffix_map.get(attr, '')} / {attr}"
                        row_data["damage_scenario_sn"] = ds.get("damage_scenario_sn", "")
                        row_data["damage_scenario"] = ds.get("damage_scenario", "")
                        
                        impact_rat = ds.get("impact_rating", {})
                        def to_str_rating(val):
                            if isinstance(val, int):
                                return ["Negligible", "Moderate", "Major", "Severe"][min(max(val, 0), 3)]
                            if isinstance(val, str) and val.isdigit():
                                return ["Negligible", "Moderate", "Major", "Severe"][min(max(int(val), 0), 3)]
                            if isinstance(val, str) and val.strip().lower() in ["negligible", "moderate", "major", "severe"]:
                                return val.strip().capitalize()
                            return str(val) if val else "Negligible"

                        row_data["safety"] = to_str_rating(impact_rat.get("safety"))
                        row_data["financial"] = to_str_rating(impact_rat.get("financial"))
                        row_data["operational"] = to_str_rating(impact_rat.get("operational"))
                        row_data["privacy"] = to_str_rating(impact_rat.get("privacy"))
                        row_data["overall_impact"] = to_str_rating(ds.get("overall_impact"))
                        
                        row_data["threat_scenario"] = ts.get("threat_scenario", "")
                        
                        # Attack Path & Feasibility Columns
                        if desensitize:
                            row_data["attack_path"] = "攻击路径分析细节已脱敏过滤。 / Attack path details have been desensitized and filtered."
                            row_data["time_consuming"] = "N/A"
                            row_data["expertise"] = "N/A"
                            row_data["knowledge_about_toe"] = "N/A"
                            row_data["window_of_opportunity"] = "N/A"
                            row_data["equipment"] = "N/A"
                            row_data["difficulty"] = "N/A"
                        else:
                            raw_ap = ap.get("attack_path", "")
                            if isinstance(raw_ap, dict):
                                ap_lines = []
                                if raw_ap.get("entry_point"):
                                    ap_lines.append(f"入口点 / Entry Point: {raw_ap.get('entry_point')}")
                                if raw_ap.get("attack_technique"):
                                    ap_lines.append(f"技术手法 / Attack Technique: {raw_ap.get('attack_technique')}")
                                if raw_ap.get("attack_steps"):
                                    steps_list = raw_ap.get("attack_steps")
                                    if isinstance(steps_list, list):
                                        ap_lines.append("步骤 / Steps:\n" + "\n".join(steps_list))
                                    else:
                                        ap_lines.append(f"步骤 / Steps: {steps_list}")
                                row_data["attack_path"] = "\n".join(ap_lines)
                            elif isinstance(ap, dict) and "entry_point" in ap:
                                ap_lines = []
                                if ap.get("entry_point"):
                                    ap_lines.append(f"入口点 / Entry Point: {ap.get('entry_point')}")
                                if ap.get("attack_technique"):
                                    ap_lines.append(f"技术手法 / Attack Technique: {ap.get('attack_technique')}")
                                if ap.get("attack_steps"):
                                    steps_list = ap.get("attack_steps")
                                    if isinstance(steps_list, list):
                                        ap_lines.append("步骤 / Steps:\n" + "\n".join(steps_list))
                                    else:
                                        ap_lines.append(f"步骤 / Steps: {steps_list}")
                                row_data["attack_path"] = "\n".join(ap_lines)
                            else:
                                row_data["attack_path"] = str(raw_ap) if raw_ap else (ap.get("method", "") or str(ap))
                                
                            tc = ap.get("time_consuming", "")
                            exp = ap.get("expertise", "")
                            toe = ap.get("knowledge_about_toe", "")
                            win = ap.get("window_of_opportunity", "")
                            eq = ap.get("equipment", "")
                            
                            tc_map = {
                                "no_more_than_1d": "≤ 1 d", "no_more_than_1w": "≤ 1 w", "no_more_than_1m": "≤ 1 m",
                                "no_more_than_6m": "≤ 6 m", "more_than_6m": "> 6 m"
                            }
                            exp_map = {
                                "layman": "Layman", "proficient": "Proficient", "expert": "Expert", "multiple_expert": "Multiple Expert"
                            }
                            toe_map = {
                                "public": "Public", "restricted": "Restricted", "confidential": "Confidential",
                                "strictly_confidential": "Strictly Confidential"
                            }
                            win_map = {
                                "unlimited": "Unlimited", "easy": "Easy", "moderate": "Moderate", "difficult": "Difficult"
                            }
                            eq_map = {
                                "standard": "Standard", "specialized": "Specialized", "bespoke": "Bespoke",
                                "multiple_bespoke": "Multiple Bespoke"
                            }
                            row_data["time_consuming"] = tc_map.get(tc, tc)
                            row_data["expertise"] = exp_map.get(exp, exp)
                            row_data["knowledge_about_toe"] = toe_map.get(toe, toe)
                            row_data["window_of_opportunity"] = win_map.get(win, win)
                            row_data["equipment"] = eq_map.get(eq, eq)
                            
                            diff = ap.get("difficulty")
                            if diff is None:
                                t_pts = {"no_more_than_1d": 0, "no_more_than_1w": 1, "no_more_than_1m": 4, "no_more_than_6m": 17, "more_than_6m": 19}.get(tc, 1)
                                ex_pts = {"layman": 0, "proficient": 3, "expert": 6, "multiple_expert": 8}.get(exp, 3)
                                to_pts = {"public": 0, "restricted": 3, "confidential": 7, "strictly_confidential": 11}.get(toe, 3)
                                w_pts = {"unlimited": 0, "easy": 1, "moderate": 4, "difficult": 10}.get(win, 1)
                                eq_pts = {"standard": 0, "specialized": 4, "bespoke": 7, "multiple_bespoke": 9}.get(eq, 0)
                                diff = t_pts + ex_pts + to_pts + w_pts + eq_pts
                            row_data["difficulty"] = diff
                            
                        feas = ap.get("feasibility", "")
                        feas_map = {"verylow": "Very Low", "low": "Low", "medium": "Medium", "high": "High"}
                        feas_disp = feas_map.get(str(feas).lower().replace(" ", ""), feas)
                        row_data["af_level"] = feas_disp
                        
                        caf = rd.get("caf_level") or feas_disp
                        caf_disp = feas_map.get(str(caf).lower().replace(" ", ""), caf)
                        row_data["caf_level"] = caf_disp
                        
                        row_data["risk_value"] = rd.get("risk_value", "")
                        row_data["risk_treatment"] = rt_disp
                        row_data["cybersecurity_claim_id"] = rd.get("cybersecurity_claim_id", "N/A") if rt_disp in ["Share", "Retain"] else "N/A"
                        row_data["cybersecurity_claim"] = rd.get("cybersecurity_claim", "N/A") if rt_disp in ["Share", "Retain"] else "N/A"
                        
                        goal_id_val = rd.get("cybersecurity_goal_id") or rd.get("cso_id")
                        if not goal_id_val or goal_id_val == "N/A":
                            if matching_reqs and matching_reqs[0].get("cybersecurity_control_id"):
                                goal_id_val = matching_reqs[0].get("cybersecurity_control_id")
                            else:
                                goal_id_val = f"CSO_{attr}_{ts.get('threat_id')}"
                        row_data["cybersecurity_goal_id"] = goal_id_val if rt_disp == "Reduce" else "N/A"
                        row_data["cybersecurity_goal"] = rd.get("cybersecurity_goal", "N/A") if rt_disp == "Reduce" else "N/A"
                        
                        row_data["cybersecurity_control_id"] = cybersecurity_control_id_val
                        row_data["cybersecurity_control"] = cybersecurity_control_val
                        row_data["allocated_to_device"] = allocated_to_device_val
                        row_data["cybersecurity_requirement_id"] = cybersecurity_requirement_id_val
                        row_data["cybersecurity_requirement"] = cybersecurity_requirement_val
                        
                        rows.append(row_data)
                        row_num += 1
                        
    goal_map, claim_map, control_map, req_map = build_id_maps(steps)
    for row in rows:
        rt_disp = row.get("risk_treatment")
        is_exempted = rt_disp in ["Share", "Retain"]
        is_mitigated = rt_disp == "Reduce"

        # 1. Claims
        claim_val = str(row.get("cybersecurity_claim") or "").strip()
        if is_exempted and claim_val and claim_val != "N/A":
            row["cybersecurity_claim_id"] = claim_map.get(claim_val, "N/A")
        else:
            row["cybersecurity_claim_id"] = "N/A"
            row["cybersecurity_claim"] = "N/A" if (is_mitigated or rt_disp == "Avoid") else row.get("cybersecurity_claim", "N/A")

        # 2. Goals (CSO)
        goal_val = str(row.get("cybersecurity_goal") or "").strip()
        if is_mitigated and goal_val and goal_val != "N/A":
            row["cybersecurity_goal_id"] = goal_map.get(goal_val, "N/A")
        else:
            row["cybersecurity_goal_id"] = "N/A"
            row["cybersecurity_goal"] = "N/A" if (is_exempted or rt_disp == "Avoid") else row.get("cybersecurity_goal", "N/A")

        # 3. Controls
        control_val = str(row.get("cybersecurity_control") or "").strip()
        if is_mitigated and control_val and control_val != "N/A":
            lines = [line.strip() for line in control_val.split("\n") if line.strip()]
            new_ids = []
            for line in lines:
                import re
                cleaned = re.sub(r'^\(\d+\)\s*', '', line).strip()
                if cleaned and cleaned != "N/A":
                    new_ids.append(control_map.get(cleaned, "CSC-0000"))
            if len(new_ids) > 1:
                row["cybersecurity_control_id"] = "\n".join(f"({i+1}) {cid}" for i, cid in enumerate(new_ids))
            else:
                row["cybersecurity_control_id"] = new_ids[0] if new_ids else "N/A"
        else:
            row["cybersecurity_control_id"] = "N/A"

        # 4. Requirements (CSR)
        req_val = str(row.get("cybersecurity_requirement") or "").strip()
        if is_mitigated and req_val and req_val != "N/A":
            lines = [line.strip() for line in req_val.split("\n") if line.strip()]
            new_ids = []
            for line in lines:
                import re
                cleaned = re.sub(r'^\(\d+\)\s*', '', line).strip()
                if cleaned and cleaned != "N/A":
                    new_ids.append(req_map.get(cleaned, "CSR-0000"))
            if len(new_ids) > 1:
                row["cybersecurity_requirement_id"] = "\n".join(f"({i+1}) {rid}" for i, rid in enumerate(new_ids))
            else:
                row["cybersecurity_requirement_id"] = new_ids[0] if new_ids else "N/A"
        else:
            row["cybersecurity_requirement_id"] = "N/A"

    return rows

def collect_csr_report_data(steps: List[TaraStep], assets: List[Asset]) -> List[dict]:
    """
    收集并去重汇总所有资产在 Stage 5 生成的 summarized_csrs 网络安全需求列表
    """
    steps_by_asset = {}
    for step in steps:
        if step.asset_id not in steps_by_asset:
            steps_by_asset[step.asset_id] = {}
        steps_by_asset[step.asset_id][step.stage] = step
        
    rows = []
    csr_num = 1
    
    # 动态推断 security_domain 的辅助函数，兼容较早未打标 security_domain 的数据
    def infer_security_domain(req_text: str) -> str:
        req_lower = req_text.lower()
        if any(w in req_lower for w in ["ota", "update", "升级"]):
            return "安全升级 / Secure Update"
        if any(w in req_lower for w in ["secoc", "transmission", "通信", "message", "bus"]):
            return "安全通信 / Secure Transmission"
        if any(w in req_lower for w in ["crypt", "encrypt", "key", "signature", "sign", "verify", "hash", "算法", "密码", "秘钥", "密钥"]):
            return "密码学与存储安全 / Cryptography & Storage Security"
        if any(w in req_lower for w in ["diag", "diagnose", "诊断"]):
            return "诊断安全 / Diagnostics Security"
        if any(w in req_lower for w in ["access", "auth", "login", "permission", "访问", "授权", "身份"]):
            return "访问控制 / Access Control"
        return "通用安全 / General Security"

    for asset in sorted(assets, key=lambda x: x.id):
        asset_steps = steps_by_asset.get(asset.id, {})
        if not asset_steps:
            continue
            
        stage5_out = asset_steps.get("stage5").analysis_result.get("final_output", {}) if asset_steps.get("stage5") else {}
        summarized_csrs = stage5_out.get("summarized_csrs", [])
        
        # 兜底：如果 summarized_csrs 为空，从 requirements 中提取 allocated_to_device == "yes" 的项进行拼装
        if not summarized_csrs:
            requirements = stage5_out.get("requirements", [])
            device_reqs = [req for req in requirements if str(req.get("allocated_to_device", "")).lower().strip() in ["yes", "true"]]
            
            seen_req_texts = set()
            summarized_csrs = []
            for req in device_reqs:
                req_text = req.get("cybersecurity_requirement") or ""
                if not req_text or req_text in seen_req_texts:
                    continue
                seen_req_texts.add(req_text)
                
                req_id = req.get("cybersecurity_requirement_id") or f"CSR-{asset.id:03d}-{len(seen_req_texts):02d}"
                domain_val = req.get("security_domain") or infer_security_domain(req_text)
                
                summarized_csrs.append({
                    "asset_id": f"ID{asset.id}",
                    "asset_name": asset.name,
                    "cybersecurity_requirement_id": req_id,
                    "csr_id": req_id,
                    "title": f"针对 {asset.name} 的安全要求 / Security requirement for {asset.name}",
                    "sub_title": f"网络安全防护 / Cybersecurity protection for {asset.name}",
                    "security_domain": domain_val,
                    "cybersecurity_requirement": req_text
                })
                
        for csr in summarized_csrs:
            c_id = csr.get("csr_id") or csr.get("cybersecurity_requirement_id") or f"CSR_{csr_num:04d}"
            rows.append({
                "number": f"CSR_{csr_num:04d}",
                "csr_id": c_id,
                "security_domain": csr.get("security_domain") or "通用安全 / General Security",
                "asset_sn": get_asset_sn(asset),
                "asset_name": asset.name,
                "title": csr.get("title") or "N/A",
                "sub_title": csr.get("sub_title") or "N/A",
                "cybersecurity_requirement": csr.get("cybersecurity_requirement") or "N/A"
            })
            csr_num += 1
            
    _, _, _, req_map = build_id_maps(steps)
    for row in rows:
        req_text = str(row.get("cybersecurity_requirement") or "").strip()
        if req_text and req_text != "N/A":
            row["csr_id"] = req_map.get(req_text, row.get("csr_id"))
        else:
            row["csr_id"] = "N/A"

    return rows

def create_excel_report(domain: Domain, steps: List[TaraStep], assets: List[Asset], desensitize: bool, export_type: str = "all") -> str:
    """
    生成高度美观的 ISO 21434 TARA XLSX 格式报表，支持脱敏过滤，并支持导出安全需求 (CSR_Requirements) 的混合模型 sheet 页
    """
    wb = openpyxl.Workbook()
    default_ws = wb.active
    
    # 颜色与样式设定 (Dark Gray Header Theme)
    header_fill_r1 = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_fill_r2 = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
    cell_font = Font(name="Microsoft YaHei", size=9.5)
    
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    has_tara = export_type in ["all", "tara"]
    has_csr = export_type in ["all", "csr"]
    
    # 导出 TARA Sheet
    if has_tara:
        ws = default_ws
        ws.title = "TARA"
        ws.views.sheetView[0].showGridLines = True
        
        # 1. 标题行高度
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 35
        
        # 填充所有表头单元格默认格式
        for col_idx in range(1, 33):
            cell_r1 = ws.cell(row=1, column=col_idx)
            cell_r1.fill = header_fill_r1
            cell_r1.font = header_font
            cell_r1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_r1.border = thin_border
            
            cell_r2 = ws.cell(row=2, column=col_idx)
            cell_r2.fill = header_fill_r2
            cell_r2.font = header_font
            cell_r2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_r2.border = thin_border
            
        # 第一层合并表头 (Row 1)
        ws.merge_cells('B1:C1')
        ws['B1'] = 'Assets'
        ws['D1'] = 'Cybersecurity Attributes Result'
        ws.merge_cells('E1:K1')
        ws['E1'] = 'Damage Scenarios and Impact Category'
        ws.merge_cells('L1:U1')
        ws['L1'] = 'Threat Scenarios and Attack Feasibility Assessment'
        
        # 第二层标题名称 (Row 2)
        headers_r2 = [
            "Number", "Assets SN", "Assets Name",
            "Cybersecurity Attributes Result",
            "Damage Scenarios SN", "Damage Scenarios", "Safety", "Financial", "Operational", "Privacy", "Impact Level",
            "Threat Scenarios", "Attack Path",
            "Time Consuming", "Expertise", "Knowledge about TOE", "Window of opportunity", "Equipment", "Difficulty", "AF Level", "CAF Level",
            "Risk Value", "Risk Treatment Recommend",
            "Cybersecurity Claims ID", "Cybersecurity Claims", "Cybersecurity Goal ID", "Cybersecurity Goal", "Cybersecurity Control ID", "Cybersecurity Control", "Allocated to ADCU", "Cybersecurity Requirement ID", "Cybersecurity Requirement"
        ]
        
        for col_idx, h in enumerate(headers_r2, 1):
            ws.cell(row=2, column=col_idx, value=h)
            
        rows = collect_report_data(steps, assets, desensitize)
        
        current_row = 3
        keys = [
            "number", "asset_sn", "asset_name",
            "attribute_result",
            "damage_scenario_sn", "damage_scenario", "safety", "financial", "operational", "privacy", "overall_impact",
            "threat_scenario", "attack_path",
            "time_consuming", "expertise", "knowledge_about_toe", "window_of_opportunity", "equipment", "difficulty", "af_level", "caf_level",
            "risk_value", "risk_treatment",
            "cybersecurity_claim_id", "cybersecurity_claim", "cybersecurity_goal_id", "cybersecurity_goal", "cybersecurity_control_id", "cybersecurity_control", "allocated_to_device", "cybersecurity_requirement_id", "cybersecurity_requirement"
        ]
        
        for row_data in rows:
            for col_idx, key in enumerate(keys, 1):
                val = row_data.get(key, "")
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = cell_font
                cell.border = thin_border
                
                # 描述文本左对齐，其它代号及评价指标等均居中对齐
                if key in ["damage_scenario", "threat_scenario", "attack_path", "cybersecurity_requirement", "cybersecurity_control", "cybersecurity_claim", "cybersecurity_goal"]:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
            ws.row_dimensions[current_row].height = 24
            current_row += 1
                
        # 自适应列宽设定
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.row in [1, 2]:
                    continue
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 11), 35)

    # 导出 CSR Sheet (混合模式)
    if has_csr:
        if has_tara:
            ws2 = wb.create_sheet(title="CSR_Requirements")
        else:
            ws2 = default_ws
            ws2.title = "CSR_Requirements"
            
        ws2.views.sheetView[0].showGridLines = True
        
        # 填充所有表头单元格默认格式
        for col_idx in range(1, 8):
            cell_r1 = ws2.cell(row=1, column=col_idx)
            cell_r1.fill = header_fill_r1
            cell_r1.font = header_font
            cell_r1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_r1.border = thin_border
            
            cell_r2 = ws2.cell(row=2, column=col_idx)
            cell_r2.fill = header_fill_r2
            cell_r2.font = header_font
            cell_r2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_r2.border = thin_border

        ws2.row_dimensions[1].height = 30
        ws2.row_dimensions[2].height = 35
        
        # 第一层合并表头 (Row 1 for Sheet 2)
        ws2['A1'] = 'CSR ID'
        ws2['B1'] = 'Security Domain'
        ws2.merge_cells('C1:D1')
        ws2['C1'] = 'Assets Allocation'
        ws2.merge_cells('E1:G1')
        ws2['E1'] = 'Cybersecurity Requirements Details'
        
        # 第二层标题名称 (Row 2 for Sheet 2)
        headers2_r2 = [
            "CSR ID", "Security Domain", "Asset SN", "Asset Name",
            "Requirement Title", "Requirement Subtitle", "Cybersecurity Requirement"
        ]
        for col_idx, h in enumerate(headers2_r2, 1):
            ws2.cell(row=2, column=col_idx, value=h)
            
        csr_rows = collect_csr_report_data(steps, assets)
        
        current_row = 3
        csr_keys = ["csr_id", "security_domain", "asset_sn", "asset_name", "title", "sub_title", "cybersecurity_requirement"]
        
        for row_data in csr_rows:
            for col_idx, key in enumerate(csr_keys, 1):
                val = row_data.get(key, "")
                cell = ws2.cell(row=current_row, column=col_idx, value=val)
                cell.font = cell_font
                cell.border = thin_border
                
                # 描述文本左对齐，其它居中对齐
                if key in ["title", "sub_title", "cybersecurity_requirement"]:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
            ws2.row_dimensions[current_row].height = 24
            current_row += 1
            
        # 自适应列宽设定 (Sheet 2)
        for col in ws2.columns:
            max_len = 0
            for cell in col:
                if cell.row in [1, 2]:
                    continue
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws2.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)
            
    filename_suffix = f"{export_type}_{'desensitized' if desensitize else 'full'}"
    file_path = os.path.join(EXPORTS_DIR, f"TARA_Report_{domain.id}_{filename_suffix}.xlsx")
    wb.save(file_path)
    return file_path

def create_csv_report(domain: Domain, steps: List[TaraStep], assets: List[Asset], desensitize: bool, export_type: str = "tara") -> str:
    """
    生成 ISO 21434 TARA CSV 格式报表，根据 export_type 可选导出 TARA 威胁分析 或 CSR 需求列表
    使用 utf-8-sig 编码，防止中文在 Excel 中打开时显示乱码
    """
    filename_suffix = f"{export_type}_{'desensitized' if desensitize else 'full'}"
    file_path = os.path.join(EXPORTS_DIR, f"TARA_Report_{domain.id}_{filename_suffix}.csv")
    
    if export_type == "csr":
        headers_r2 = [
            "CSR ID", "Security Domain", "Asset SN", "Asset Name",
            "Requirement Title", "Requirement Subtitle", "Cybersecurity Requirement"
        ]
        keys = ["csr_id", "security_domain", "asset_sn", "asset_name", "title", "sub_title", "cybersecurity_requirement"]
        rows = collect_csr_report_data(steps, assets)
    else:
        headers_r2 = [
            "Number", "Assets SN", "Assets Name",
            "Cybersecurity Attributes Result",
            "Damage Scenarios SN", "Damage Scenarios", "Safety", "Financial", "Operational", "Privacy", "Impact Level",
            "Threat Scenarios", "Attack Path",
            "Time Consuming", "Expertise", "Knowledge about TOE", "Window of opportunity", "Equipment", "Difficulty", "AF Level", "CAF Level",
            "Risk Value", "Risk Treatment Recommend",
            "Cybersecurity Claims ID", "Cybersecurity Claims", "Cybersecurity Goal ID", "Cybersecurity Goal", "Cybersecurity Control ID", "Cybersecurity Control", "Allocated to ADCU", "Cybersecurity Requirement ID", "Cybersecurity Requirement"
        ]
        keys = [
            "number", "asset_sn", "asset_name",
            "attribute_result",
            "damage_scenario_sn", "damage_scenario", "safety", "financial", "operational", "privacy", "overall_impact",
            "threat_scenario", "attack_path",
            "time_consuming", "expertise", "knowledge_about_toe", "window_of_opportunity", "equipment", "difficulty", "af_level", "caf_level",
            "risk_value", "risk_treatment",
            "cybersecurity_claim_id", "cybersecurity_claim", "cybersecurity_goal_id", "cybersecurity_goal", "cybersecurity_control_id", "cybersecurity_control", "allocated_to_device", "cybersecurity_requirement_id", "cybersecurity_requirement"
        ]
        rows = collect_report_data(steps, assets, desensitize)
    
    with open(file_path, mode="w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers_r2)
        for row in rows:
            writer.writerow([row.get(k, "") if row.get(k) is not None else "" for k in keys])
            
    return file_path

# ----------------- 导出报告接口 -----------------

@router.get("/domains/{domain_id}/export")
def export_report(
    domain_id: int,
    format: str = "xlsx", # xlsx 或 csv
    desensitize: bool = False, # 是否脱敏导出 (BR-57, BR-77)
    export_type: str = "all", # all (导出全部Sheets，仅对xlsx有效), tara (只导出威胁分析), csr (只导出安全需求汇总)
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    导出网络安全 TARA 评估报告与 CSR 安全需求汇总 (XLSX/CSV，支持数据脱敏与混合模型导出，BR-57, BR-77)
    """
    domain = db.query(Domain).filter(Domain.id == domain_id).first()
    if not domain:
        raise HTTPException(status_code=404, detail="子域控不存在")
        
    # 获取最新的已完成运行记录和步骤结果
    last_run = db.query(TaraRun).filter(
        TaraRun.domain_id == domain_id,
        TaraRun.status == "completed"
    ).order_by(TaraRun.id.desc()).first()
    
    if not last_run:
        raise HTTPException(status_code=400, detail="该域控当前没有已成功完成 of TARA 报告记录，无法执行导出。")
        
    steps = db.query(TaraStep).filter(TaraStep.run_id == last_run.id).all()
    assets = db.query(Asset).filter(Asset.domain_id == domain_id).all()
    
    export_type_lower = export_type.lower().strip()
    if format.lower() == "csv" and export_type_lower == "all":
        # CSV不支持多sheet，默认设置为 tara 导出
        export_type_lower = "tara"
        
    if export_type_lower not in ["all", "tara", "csr"]:
        raise HTTPException(status_code=400, detail="不支持的导出类型，可选值: all, tara, csr。")
    
    if format.lower() == "xlsx":
        file_path = create_excel_report(domain, steps, assets, desensitize, export_type_lower)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"TARA_Report_{domain.name}_{export_type_lower}_{'desensitized' if desensitize else 'full'}.xlsx"
    elif format.lower() == "csv":
        file_path = create_csv_report(domain, steps, assets, desensitize, export_type_lower)
        media_type = "text/csv; charset=utf-8-sig"
        filename = f"TARA_Report_{domain.name}_{export_type_lower}_{'desensitized' if desensitize else 'full'}.csv"
    else:
        raise HTTPException(status_code=400, detail="不支持的导出文件格式，仅支持 'xlsx' 或 'csv'。")
        
    return FileResponse(
        path=file_path,
        media_type=media_type,
        filename=filename
    )
